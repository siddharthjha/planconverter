import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook, Workbook

st.title("Model Rollout Aggregator")

def load_excel_as_values(uploaded_file):
    wb = load_workbook(uploaded_file, data_only=True)
    sheet = wb.active

    # Create a new workbook to hold only values
    new_wb = Workbook()
    new_ws = new_wb.active

    for row in sheet.iter_rows():
        new_ws.append([cell.value for cell in row])

    # Save to BytesIO
    temp = BytesIO()
    new_wb.save(temp)
    temp.seek(0)
    return pd.read_excel(temp, header=None)

uploaded_file = st.file_uploader("Upload Input Excel File", type=["xlsx"])

if uploaded_file:
    # Step 0: Paste as values
    df_raw = load_excel_as_values(uploaded_file)

    # Step 0.5: Drop first row (to remove problematic header)
    df_raw = df_raw.iloc[1:].reset_index(drop=True)

    # Step 1: Show raw data after drop
    st.subheader("Raw Data After Removing First Row")
    st.dataframe(df_raw.head(20))

    # Step 2: Forward-fill column A (Model)
    df_raw.iloc[:, 0] = df_raw.iloc[:, 0].ffill()

    # Step 3: Filter where column C = "Rollout"
    df_rollout = df_raw[df_raw.iloc[:, 2].astype(str).str.strip().str.lower() == "rollout"].copy()
    st.subheader("Filtered Rollout Rows")
    st.dataframe(df_rollout)

    # Step 4: Aggregate data using mapped models and dates
    model_mapping = {
        "EX 200 Infra Super Plus": "EX200",
        "EX 200 Prime": "EX200",
        "EX 210 Infra Super Plus": "EX200",
        "EX 210 Prime": "EX200",
        "EX 215 Prime": "EX200",
        "ZX 220 GI Ultra": "ZX220GI",
        "ZX 220 GI (Export)": "ZX220EX",
        "EX 350LC": "ZX350LC",
        "ZX 370 GI": "ZX370",
        "ZX 370 Ultra": "ZX370",
        "ZX 400 GI": "ZX370",
        "ZX 490 Ultra": "ZX490",
        "EX 218": "ZX218"
    }

    # Detect date columns from new row 0 (after original first row is dropped)
    date_row = df_raw.iloc[0, 5:]
    date_cols = {
        col_idx: parsed.date()
        for col_idx in range(5, len(df_raw.columns))
        if pd.notna((parsed := pd.to_datetime(df_raw.iloc[0, col_idx], errors="coerce"))) and isinstance(parsed, pd.Timestamp)
    }

    output_models = sorted(set(model_mapping.values()))
    output_dates = sorted(set(date_cols.values()))
    result_df = pd.DataFrame(0, index=output_dates, columns=output_models)

    for idx, row in df_rollout.iterrows():
        model_raw = str(row[0]).strip()
        model_mapped = model_mapping.get(model_raw)
        if model_mapped:
            for col_idx, date_val in date_cols.items():
                val = row[col_idx]
                if isinstance(val, (int, float)) and pd.notna(val) and 0 <= val <= 100:
                    result_df.at[date_val, model_mapped] += val

    # Step 5: Final output formatting
    desired_order = ["EX200", "ZX220EX", "ZX220GI", "ZX350LC", "ZX370", "ZX490", "ZX218"]
    # filtered_result_df = result_df[result_df.index != pd.to_datetime("1970-01-01").date()]

    # # Reorder columns and sort by date descending
    # filtered_result_df = filtered_result_df.reindex(columns=desired_order)
    # filtered_result_df = filtered_result_df.sort_index(ascending=False)

    # Remove bad rows like 1970
    filtered_result_df = result_df[result_df.index != pd.to_datetime("1970-01-01").date()]

    # Fill missing dates between min and max
    if not filtered_result_df.empty:
        min_date, max_date = filtered_result_df.index.min(), filtered_result_df.index.max()
        full_date_range = pd.date_range(start=min_date, end=max_date, freq="D").date

        # Reindex to include all dates and fill missing rows with 0
        filtered_result_df = filtered_result_df.reindex(full_date_range, fill_value=0)

    # Reorder columns
    filtered_result_df = filtered_result_df.reindex(columns=desired_order)
    filtered_result_df = filtered_result_df.sort_index(ascending=False)

    # Format date and reset index
    formatted_result_df = filtered_result_df.copy()
    formatted_result_df.index = filtered_result_df.index.map(lambda d: d.strftime("%d/%m/%Y"))
    formatted_result_df.reset_index(inplace=True)
    formatted_result_df.rename(columns={"index": "Date"}, inplace=True)

    # Show formatted table
    st.subheader("Final Formatted Output Table")
    st.dataframe(formatted_result_df, use_container_width=True, hide_index=True)

    st.subheader("Aggregated Result Table")
    st.dataframe(result_df)

    # Optional: Download button
    def to_excel(df):
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="daily")
        return output.getvalue()

    st.download_button(
        label="Download Final Formatted Excel",
        data=to_excel(formatted_result_df),
        file_name="final_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    import sys
    if getattr(sys, 'frozen', False):
        input("Press Enter to exit...")


    # st.download_button(
    #     label="Download Aggregated Excel",
    #     data=to_excel(result_df),
    #     file_name="aggregated_output.xlsx",
    #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # )